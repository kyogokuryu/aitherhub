"""
Excel parser for clean video uploads.
Parses product.xlsx and trend_stats.xlsx files
and returns structured data for report generation.
"""
import os
import logging
import requests

logger = logging.getLogger("process_video")


def download_excel(blob_url: str, dest_path: str) -> bool:
    """Download an Excel file from Azure Blob URL."""
    if not blob_url:
        return False
    try:
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with requests.get(blob_url, stream=True, timeout=60) as r:
            r.raise_for_status()
            with open(dest_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
        logger.info(f"[EXCEL] Downloaded: {dest_path}")
        return True
    except Exception as e:
        logger.warning(f"[EXCEL] Download failed: {e}")
        return False


def parse_product_excel(file_path: str) -> list[dict]:
    """
    Parse product.xlsx.
    Expected columns (flexible matching):
    - 商品名 / product_name / name
    - 価格 / price
    - カテゴリ / category
    - その他の列も取り込む
    Returns list of product dicts.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("[EXCEL] openpyxl not installed, cannot parse Excel")
        return []

    if not os.path.exists(file_path):
        return []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        # First row = headers
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        products = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            product = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    product[headers[i]] = val
            products.append(product)

        wb.close()
        logger.info(f"[EXCEL] Parsed {len(products)} products from {file_path}")
        return products

    except Exception as e:
        logger.warning(f"[EXCEL] Failed to parse product Excel: {e}")
        return []


def parse_trend_excel(file_path: str) -> list[dict]:
    """
    Parse trend_stats.xlsx.
    Expected columns (flexible matching):
    - 時間 / time / timestamp
    - 売上 / sales / revenue
    - 注文数 / orders
    - 商品名 / product_name
    - その他の列も取り込む
    Returns list of trend data dicts.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("[EXCEL] openpyxl not installed, cannot parse Excel")
        return []

    if not os.path.exists(file_path):
        return []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        trend_data = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            entry = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    entry[headers[i]] = val
            trend_data.append(entry)

        wb.close()
        logger.info(f"[EXCEL] Parsed {len(trend_data)} trend entries from {file_path}")
        return trend_data

    except Exception as e:
        logger.warning(f"[EXCEL] Failed to parse trend Excel: {e}")
        return []


def load_excel_data(video_id: str, excel_urls: dict, work_dir: str = "excel_data") -> dict:
    """
    Download and parse both Excel files for a video.
    Returns dict with 'products' and 'trends' keys.
    """
    excel_dir = os.path.join(work_dir, video_id)
    os.makedirs(excel_dir, exist_ok=True)

    result = {
        "products": [],
        "trends": [],
        "has_product_data": False,
        "has_trend_data": False,
    }

    # Download and parse product data
    product_url = excel_urls.get("excel_product_blob_url")
    if product_url:
        product_path = os.path.join(excel_dir, "product.xlsx")
        if download_excel(product_url, product_path):
            result["products"] = parse_product_excel(product_path)
            result["has_product_data"] = len(result["products"]) > 0

    # Download and parse trend data
    trend_url = excel_urls.get("excel_trend_blob_url")
    if trend_url:
        trend_path = os.path.join(excel_dir, "trend_stats.xlsx")
        if download_excel(trend_url, trend_path):
            result["trends"] = parse_trend_excel(trend_path)
            result["has_trend_data"] = len(result["trends"]) > 0

    logger.info(
        f"[EXCEL] Loaded data for {video_id}: "
        f"{len(result['products'])} products, "
        f"{len(result['trends'])} trend entries"
    )

    return result


def format_excel_data_for_prompt(excel_data: dict) -> str:
    """
    Format Excel data into a text summary for GPT prompts.
    """
    parts = []

    if excel_data.get("has_product_data"):
        parts.append("【商品データ】")
        for i, p in enumerate(excel_data["products"][:20], 1):  # Max 20 products
            items = [f"{k}: {v}" for k, v in p.items() if v is not None]
            parts.append(f"  {i}. " + " / ".join(items))

    if excel_data.get("has_trend_data"):
        parts.append("\n【売上トレンドデータ】")
        for i, t in enumerate(excel_data["trends"][:50], 1):  # Max 50 entries
            items = [f"{k}: {v}" for k, v in t.items() if v is not None]
            parts.append(f"  {i}. " + " / ".join(items))

    return "\n".join(parts) if parts else ""


def match_sales_to_phase(trends: list[dict], start_sec: float, end_sec: float) -> dict:
    """
    Match trend/sales data to a specific phase time range.
    Returns aggregated sales metrics for the phase.

    Tries to match using time-based columns.
    """
    import re
    from datetime import datetime, timedelta

    if not trends:
        return {"sales": None, "orders": None, "products_sold": []}

    # Detect time column
    time_keys = []
    sales_keys = []
    order_keys = []
    product_keys = []

    sample = trends[0]
    for k in sample.keys():
        kl = k.lower()
        if any(w in kl for w in ["時間", "time", "timestamp", "秒", "sec", "minute"]):
            time_keys.append(k)
        if any(w in kl for w in ["売上", "sales", "revenue", "金額", "amount"]):
            sales_keys.append(k)
        if any(w in kl for w in ["注文", "order", "件数", "count"]):
            order_keys.append(k)
        if any(w in kl for w in ["商品", "product", "item", "名前", "name"]):
            product_keys.append(k)

    phase_sales = 0
    phase_orders = 0
    products_sold = []

    for t in trends:
        # Try to extract time in seconds from the entry
        entry_time = None
        for tk in time_keys:
            val = t.get(tk)
            if val is None:
                continue
            try:
                # Try direct seconds
                entry_time = float(val)
                break
            except (ValueError, TypeError):
                pass
            try:
                # Try MM:SS or HH:MM:SS format
                val_str = str(val)
                parts = val_str.split(":")
                if len(parts) == 2:
                    entry_time = int(parts[0]) * 60 + int(parts[1])
                    break
                elif len(parts) == 3:
                    entry_time = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                    break
            except (ValueError, TypeError):
                pass

        if entry_time is None:
            continue

        # Check if this entry falls within the phase
        if start_sec <= entry_time <= end_sec:
            for sk in sales_keys:
                try:
                    phase_sales += float(t.get(sk, 0) or 0)
                except (ValueError, TypeError):
                    pass
            for ok in order_keys:
                try:
                    phase_orders += int(t.get(ok, 0) or 0)
                except (ValueError, TypeError):
                    pass
            for pk in product_keys:
                pname = t.get(pk)
                if pname and str(pname).strip():
                    products_sold.append(str(pname).strip())

    return {
        "sales": phase_sales if phase_sales > 0 else None,
        "orders": phase_orders if phase_orders > 0 else None,
        "products_sold": list(set(products_sold)),
    }
